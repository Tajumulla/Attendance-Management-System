import cv2
import tkinter as tk
from tkinter import messagebox, ttk
import face_recognition
import os
import openpyxl
from datetime import datetime

# Step 2: Face Registration
students_database = {}
model = None

def capture_images(student_id, student_name):
    face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
    video_capture = cv2.VideoCapture(0)

    # Create a directory to store images
    image_dir = f'student_images/{student_id}'
    os.makedirs(image_dir, exist_ok=True)

    image_count = 0

    while image_count < 30:  # Capture 30 images for training
        ret, frame = video_capture.read()
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

        for (x, y, w, h) in faces:
            face = frame[y:y + h, x:x + w]
            cv2.imwrite(f'{image_dir}/image_{image_count}.jpg', face)
            image_count += 1

        cv2.imshow('Capture Faces', frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    video_capture.release()
    cv2.destroyAllWindows()

# Step 3: Facial Recognition
def train_face_recognition_model():
    global model
    face_encodings = []
    labels = []
    students_ids = list(students_database.keys())

    for student_id in students_ids:
        image_dir = f'student_images/{student_id}'
        for filename in os.listdir(image_dir):
            path = os.path.join(image_dir, filename)
            image = face_recognition.load_image_file(path)

            # Use face_recognition to find face encodings
            face_encodings_in_image = face_recognition.face_encodings(image)

            # Check if at least one face is detected
            if len(face_encodings_in_image) > 0:
                encoding = face_encodings_in_image[0]
                face_encodings.append(encoding)
                labels.append(str(student_id))

            else:
                print(f"No face detected in image: {path}")

    # Create and save the model
    model = {'encodings': face_encodings, 'labels': labels}
    messagebox.showinfo("Message", "Face recognition model trained successfully.")

# Step 4: Attendance Records
class_session = 1

def mark_attendance():
    global model
    video_capture = cv2.VideoCapture(0)
    attendance_data = {}

    while True:
        ret, frame = video_capture.read()
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

        # Use face_recognition library to find faces in the frame
        face_locations = face_recognition.face_locations(frame)
        face_encodings = face_recognition.face_encodings(frame, face_locations)

        for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
            # Compare the face encoding with the saved encodings in the model
            matches = face_recognition.compare_faces(model['encodings'], face_encoding)
            name = "Unknown"

            # If a match is found, use the corresponding student ID as the name
            if True in matches:
                first_match_index = matches.index(True)
                name = str(model['labels'][first_match_index])

                # Record attendance
                if name not in attendance_data:
                    attendance_data[name] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Draw rectangle and name on the frame
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
            font = cv2.FONT_HERSHEY_DUPLEX
            cv2.putText(frame, name, (left + 6, bottom - 6), font, 0.5, (255, 255, 255), 1)

        cv2.imshow('Attendance Marking', frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    video_capture.release()
    cv2.destroyAllWindows()

    return attendance_data

# Step 5: Generate Excel Sheet
def generate_excel_sheet(attendance_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Sheet"

    # Write header row
    ws['A1'] = 'Student ID'
    ws['B1'] = 'Student Name'
    ws['C1'] = 'Attendance Time'

    # Write attendance data
    for student_id, timestamp in attendance_data.items():
        student_name = students_database.get(student_id, "Unknown")
        ws.append([str(student_id), student_name, timestamp])  # Convert student_id to string

    # Save the workbook
    excel_file_name = f'attendance_sheet_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
    wb.save(excel_file_name)

    print(f'Attendance sheet "{excel_file_name}" generated successfully.')

# Step 6: User Interface (Optional)
class AttendanceApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Attendance Management System")
        self.root.geometry("600x400")

        # Set ttkthemes style
        self.style = ttk.Style(self.root)
        self.style.theme_use("clam")  # Use the "clam" theme

        # Configure custom button style with rounded corners
        self.style.configure("Custom.TButton", padding=5, relief=tk.GROOVE, background="#64a7c8", foreground="white", borderwidth=5)
        self.style.map("Custom.TButton",
                       foreground=[('active', 'black')],
                       background=[('active', '#8bc8f1')])

        self.label = ttk.Label(self.root, text="Welcome to Attendance Management System", font=('Helvetica', 14, 'bold'))
        self.label.pack(pady=20)

        self.register_button = ttk.Button(self.root, text="Register Faces", command=self.register_faces, style="Custom.TButton")
        self.register_button.pack(pady=10)

        self.train_button = ttk.Button(self.root, text="Train Face Recognition Model", command=self.train_model, style="Custom.TButton")
        self.train_button.pack(pady=10)

        self.start_button = ttk.Button(self.root, text="Start Attendance", command=self.start_attendance, style="Custom.TButton")
        self.start_button.pack(pady=10)

        self.feedback_button = ttk.Button(self.root, text="Provide Feedback", command=self.provide_feedback, style="Custom.TButton")
        self.feedback_button.pack(pady=10)

        self.exit_button = ttk.Button(self.root, text="Exit", command=self.root.destroy, style="Custom.TButton")
        self.exit_button.pack(pady=10)

        self.footer_label = tk.Label(self.root, text="Made by Tajju", font=('Arial', 10), fg='blue', cursor='hand2')
        self.footer_label.pack(side=tk.BOTTOM, pady=10)
        self.footer_label.bind("<Button-1>", self.open_email)

    def open_email(self, event):
        import webbrowser
        webbrowser.open("mailto:tajju3069@gmail.com")

    def provide_feedback(self):
        feedback_window = tk.Toplevel(self.root)
        feedback_window.title("Feedback Form")

        feedback_label = tk.Label(feedback_window, text="Provide your feedback:")
        feedback_label.pack(pady=10)

        feedback_text = tk.Text(feedback_window, height=5, width=50)
        feedback_text.pack(pady=10)

        submit_button = ttk.Button(feedback_window, text="Submit", command=lambda: self.submit_feedback(feedback_text.get("1.0", tk.END)))
        submit_button.pack(pady=10)

    def submit_feedback(self, feedback):
        messagebox.showinfo("Feedback Submitted", "Thank you for providing feedback!")
        
    def register_faces(self):
        student_id = input("Enter Student ID: ")
        student_name = input("Enter Student Name: ")
        students_database[student_id] = student_name
        capture_images(student_id, student_name)
        messagebox.showinfo("Message", f"Faces registered for Student {student_id}: {student_name}")

    def train_model(self):
        train_face_recognition_model()

    def start_attendance(self):
        messagebox.showinfo("Message", "Starting attendance session!")
        attendance_data = mark_attendance()
        generate_excel_sheet(attendance_data)
        messagebox.showinfo("Message", "Attendance session completed and Excel sheet generated!")

    def run(self):
        self.root.mainloop()

# Example usage
attendance_app = AttendanceApp()
attendance_app.run()
